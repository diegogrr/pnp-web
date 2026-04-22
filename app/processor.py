import io

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows

import config

# ---------------------------------------------------------------------------
# Constantes extraídas do notebook v2 — NÃO ALTERAR
# ---------------------------------------------------------------------------

ano_inicial = 2019
ano_final = 2024

dic_ordem = {
    'Quant.': ['Quant.', '%'],
    'Cor / Raça': ['Amarela', 'Branca', 'Indígena', 'Parda', 'Preta', 'Não declarada'],
    'Renda Familiar': ['0<RFP<=0,5', '0,5<RFP<=1', '1<RFP<=1,5', '1,5<RFP<=2,5', '2,5<RFP<=3,5', 'RFP>3,5', 'Não declarada'],
    'Sexo': ['M', 'F'],
    'Faixa Etária': ['Menor de 14 anos',
          '15 a 19 anos',
          '20 a 24 anos',
          '25 a 29 anos',
          '30 a 34 anos',
          '35 a 39 anos',
          '40 a 44 anos',
          '45 a 49 anos',
          '50 a 54 anos',
          '55 a 59 anos',
          'Maior de 60 anos']
}

# Mapeamento de Categoria da Situação -> Situação (notebook v2, cell [46])
DIC_SITUACAO = {
    'Concluintes': 'Conclusão',
    'Concluídos':  'Conclusão',
    'Evadidos':    'Evasão',
    'Em Curso':    'Retenção',
    'Em curso':    'Retenção',
}

# ---------------------------------------------------------------------------
# Funções extraídas do notebook v2 (cell [31]) — NÃO ALTERAR
# ---------------------------------------------------------------------------

def gera_tabela_estratificada(df, prop, df_denominador=None):
    df_estrat = df.groupby(['Ano', prop])['Código da Matricula'].count().to_frame(name='Quant.')

    if df_denominador is None:
        # Matrícula: % = subcategoria / total do ano (comportamento original)
        total_ano = df_estrat.groupby(level=0).transform('sum')
    else:
        # Conclusão/Evasão/Retenção: % = subcategoria / total da mesma
        # subcategoria no df_denominador (ex: brancos concl. / brancos total)
        total_ano = df_denominador.groupby(['Ano', prop])['Código da Matricula'].count().to_frame(name='Quant.')

    df_estrat['%'] = (df_estrat['Quant.']/total_ano['Quant.']*100).round(2)

    df_estrat = df_estrat.unstack().swaplevel(0, 1, axis=1)

    index_row = pd.Index(range(ano_inicial, ano_final+1))
    df_estrat = df_estrat.reindex(index_row)

    index_col = pd.MultiIndex.from_product([dic_ordem[prop], dic_ordem['Quant.']], names=[prop, 'Quantidade'])
    df_estrat = df_estrat.reindex(index_col, axis=1).convert_dtypes().fillna(0)

    return df_estrat

# ---------------------------------------------------------------------------
# Função extraída do notebook v2 (cell [32]) — NÃO ALTERAR
# ---------------------------------------------------------------------------

def escreve_tabela(wb, df, aba, col_inicio, linha_inicio):

    ws = wb[aba]

    if isinstance(col_inicio, str):
        c_idx_start = column_index_from_string(col_inicio)
    else:
        c_idx_start = col_inicio

    rows = dataframe_to_rows(df, index=False, header=False)

    for r_idx, row in enumerate(rows, start=linha_inicio):
        for c_idx, value in enumerate(row, start=c_idx_start):
            ws.cell(row=r_idx, column=c_idx, value=value)

# ---------------------------------------------------------------------------
# Funções de eficiência extraídas do notebook v2 (cells [79] e [80])
# Bug do notebook corrigido: gera_eficiencia_ciclo usa o parâmetro df_ef
# em vez da variável de escopo externo df_ef_curso.
# Guarda defensiva re-adicionada (ausente no notebook v2, necessária para
# cursos sem concluintes / evadidos / retidos no período).
# ---------------------------------------------------------------------------

def gera_eficiencia_ciclo(df_ef):
    """Gera tabela de IEA a partir do DataFrame de eficiência já mapeado."""
    df_indicadores = df_ef.groupby(['Ano', 'Situação'])['Código da Matricula'].count().unstack()

    for col in ['Conclusão', 'Retenção', 'Evasão']:
        if col not in df_indicadores.columns:
            df_indicadores[col] = 0

    df_indicadores = df_indicadores.fillna(0)
    df_indicadores['Total'] = df_indicadores.sum(axis=1)
    df_indicadores['IEA'] = ((
        df_indicadores['Conclusão'] +
        df_indicadores['Conclusão'] * df_indicadores['Retenção'] /
        (df_indicadores['Conclusão'] + df_indicadores['Evasão'])
    ) / df_indicadores['Total'] * 100).round(2)

    index_row = pd.Index(range(ano_inicial, ano_final + 1))
    df_indicadores = df_indicadores.reindex(index_row).fillna('')

    return df_indicadores


def gera_eficiencia_ciclo_estratificado(df_ef, prop):
    """Gera tabela de IEA estratificada por propriedade demográfica."""
    df_indicadores = df_ef.groupby(['Ano', prop, 'Situação'])['Código da Matricula'].count().unstack()

    for col in ['Conclusão', 'Retenção', 'Evasão']:
        if col not in df_indicadores.columns:
            df_indicadores[col] = 0

    df_indicadores['Total'] = df_indicadores.sum(axis=1)
    df_indicadores = df_indicadores.fillna(0)
    df_indicadores['IEA'] = ((
        df_indicadores['Conclusão'] +
        df_indicadores['Conclusão'] * df_indicadores['Retenção'] /
        (df_indicadores['Conclusão'] + df_indicadores['Evasão'])
    ) / df_indicadores['Total'] * 100).round(2)

    df_indicadores = df_indicadores[['IEA']].unstack()
    df_indicadores.columns = df_indicadores.columns.droplevel(0)

    index_row = pd.Index(range(ano_inicial, ano_final + 1))
    df_indicadores = df_indicadores.reindex(index_row)
    index_col = pd.Index(dic_ordem[prop])
    df_indicadores = df_indicadores.reindex(index_col, axis=1).fillna('')

    return df_indicadores

# ---------------------------------------------------------------------------
# Orquestradoras: processar / processar_multi / _processar_df
# ---------------------------------------------------------------------------

def processar(unidade, curso):
    """Executa o pipeline completo para unidade+curso (compatibilidade legado)."""
    df1    = pd.read_parquet(config.PARQUET_PATH)
    df1_ef = pd.read_parquet(config.PARQUET_EF_PATH)

    df_curso    = df1[(df1['Unidade de Ensino'] == unidade) & (df1['Nome de Curso'] == curso)].copy()
    df_ef_curso = df1_ef[(df1_ef['Unidade de Ensino'] == unidade) & (df1_ef['Nome de Curso'] == curso)].copy()

    return _processar_df(df_curso, df_ef_curso, curso)


def processar_multi(df_filtrado, df_ef_filtrado, rotulo, turno_str=''):
    """Entrada para filtros multi-select: recebe ambos os DataFrames pré-filtrados."""
    return _processar_df(df_filtrado.copy(), df_ef_filtrado.copy(), rotulo, turno_str)


def _processar_df(df_curso, df_ef_curso, rotulo, turno_str=''):
    """Pipeline interno de geração de tabelas — lógica do notebook v2."""

    # Mapeamento de Situação no DataFrame de eficiência (notebook v2, cell [46])
    df_ef_curso = df_ef_curso.copy()
    df_ef_curso['Situação'] = df_ef_curso['Categoria da Situação'].map(DIC_SITUACAO)

    index_row = pd.Index(range(ano_inicial, ano_final + 1))

    # --- Matrícula (df_curso — parquet principal) ---
    df_totais = df_curso.groupby(['Ano'])['Código da Matricula'].count().to_frame(name='Quant.')
    tabela1 = df_totais.reindex(index_row).fillna(0).astype(int)
    tabela2 = gera_tabela_estratificada(df_curso, 'Cor / Raça')
    tabela3 = gera_tabela_estratificada(df_curso, 'Renda Familiar')
    tabela4 = gera_tabela_estratificada(df_curso, 'Sexo')
    tabela5 = gera_tabela_estratificada(df_curso, 'Faixa Etária')

    # Denominador de eficiência (total registros por ano no parquet de eficiência)
    df_totais_ef = df_ef_curso.groupby(['Ano'])['Código da Matricula'].count().to_frame(name='Quant.')
    df_totais_ef = df_totais_ef.reindex(index_row).fillna(0).astype(int)

    # --- Conclusão (df_ef_curso — parquet de eficiência) ---
    df_concluintes = df_ef_curso[df_ef_curso['Situação'] == 'Conclusão']
    df_totais_concl = df_concluintes.groupby(['Ano'])['Código da Matricula'].count().to_frame(name='Quant.')
    df_totais_concl = df_totais_concl.reindex(index_row)
    df_totais_concl['%'] = (df_totais_concl['Quant.'] / df_totais_ef['Quant.'] * 100).round(2)
    tabela6  = df_totais_concl.convert_dtypes().fillna(0)
    tabela7  = gera_tabela_estratificada(df_concluintes, 'Cor / Raça', df_ef_curso)
    tabela8  = gera_tabela_estratificada(df_concluintes, 'Renda Familiar', df_ef_curso)
    tabela9  = gera_tabela_estratificada(df_concluintes, 'Sexo', df_ef_curso)
    tabela10 = gera_tabela_estratificada(df_concluintes, 'Faixa Etária', df_ef_curso)

    # --- Evasão (df_ef_curso — parquet de eficiência) ---
    df_evadidos = df_ef_curso[df_ef_curso['Situação'] == 'Evasão']
    df_totais_evad = df_evadidos.groupby(['Ano'])['Código da Matricula'].count().to_frame(name='Quant.')
    df_totais_evad['%'] = (df_totais_evad['Quant.'] / df_totais_ef['Quant.'] * 100).round(2)
    tabela11 = df_totais_evad.reindex(index_row).convert_dtypes().fillna(0)
    tabela12 = gera_tabela_estratificada(df_evadidos, 'Cor / Raça', df_ef_curso)
    tabela13 = gera_tabela_estratificada(df_evadidos, 'Renda Familiar', df_ef_curso)
    tabela14 = gera_tabela_estratificada(df_evadidos, 'Sexo', df_ef_curso)
    tabela15 = gera_tabela_estratificada(df_evadidos, 'Faixa Etária', df_ef_curso)

    # --- Retenção (df_ef_curso — parquet de eficiência) ---
    df_retidos = df_ef_curso[df_ef_curso['Situação'] == 'Retenção']
    df_totais_ret = df_retidos.groupby(['Ano'])['Código da Matricula'].count().to_frame(name='Quant.')
    df_totais_ret['%'] = (df_totais_ret['Quant.'] / df_totais_ef['Quant.'] * 100).round(2)
    tabela16 = df_totais_ret.reindex(index_row).convert_dtypes().fillna(0)
    tabela17 = gera_tabela_estratificada(df_retidos, 'Cor / Raça', df_ef_curso)
    tabela18 = gera_tabela_estratificada(df_retidos, 'Renda Familiar', df_ef_curso)
    tabela19 = gera_tabela_estratificada(df_retidos, 'Sexo', df_ef_curso)
    tabela20 = gera_tabela_estratificada(df_retidos, 'Faixa Etária', df_ef_curso)

    # --- Eficiência (df_ef_curso — parquet de eficiência) ---
    tabela21 = gera_eficiencia_ciclo(df_ef_curso)[['IEA']].fillna('')
    tabela22 = gera_eficiencia_ciclo_estratificado(df_ef_curso, 'Cor / Raça')
    tabela23 = gera_eficiencia_ciclo_estratificado(df_ef_curso, 'Renda Familiar')
    tabela24 = gera_eficiencia_ciclo_estratificado(df_ef_curso, 'Sexo')
    tabela25 = gera_eficiencia_ciclo_estratificado(df_ef_curso, 'Faixa Etária')

    def _concat(*dfs):
        """Concatena DataFrames horizontalmente para visualização no frontend."""
        return pd.concat([df.reset_index() for df in dfs], axis=1)

    return {
        'Acesso': {
            'tabelas': [],
            'df': pd.DataFrame({'Curso': [rotulo], 'Turno': [turno_str]}),
        },
        'Matrícula': {
            'tabelas': [
                ('E', 7, tabela1), ('G', 7, tabela2), ('T', 7, tabela3),
                ('AI', 7, tabela4), ('AN', 7, tabela5),
            ],
            'df': _concat(tabela1, tabela2, tabela3, tabela4, tabela5),
        },
        'Conclusão': {
            'tabelas': [
                ('E', 7, tabela6), ('H', 7, tabela7), ('U', 7, tabela8),
                ('AJ', 7, tabela9), ('AO', 7, tabela10),
            ],
            'df': _concat(tabela6, tabela7, tabela8, tabela9, tabela10),
        },
        'Evasão': {
            'tabelas': [
                ('E', 7, tabela11), ('H', 7, tabela12), ('U', 7, tabela13),
                ('AJ', 7, tabela14), ('AO', 7, tabela15),
            ],
            'df': _concat(tabela11, tabela12, tabela13, tabela14, tabela15),
        },
        'Retenção': {
            'tabelas': [
                ('E', 7, tabela16), ('H', 7, tabela17), ('U', 7, tabela18),
                ('AJ', 7, tabela19), ('AO', 7, tabela20),
            ],
            'df': _concat(tabela16, tabela17, tabela18, tabela19, tabela20),
        },
        'Eficiência': {
            'tabelas': [
                ('E', 7, tabela21), ('G', 7, tabela22), ('N', 7, tabela23),
                ('V', 7, tabela24), ('Y', 7, tabela25),
            ],
            'df': _concat(tabela21, tabela22, tabela23, tabela24, tabela25),
        },
        'PAP': {
            'tabelas': [],
            'df': pd.DataFrame(),
        },
    }

# ---------------------------------------------------------------------------
# Orquestradora: exportar_xlsx
# ---------------------------------------------------------------------------

def exportar_xlsx(dados, nome_arquivo):
    """Escreve dados na planilha-modelo e retorna bytes do .xlsx."""
    with open(config.MODELO_PATH, 'rb') as f:
        buf = io.BytesIO(f.read())
    wb = load_workbook(buf)

    rotulo = dados['Acesso']['df']['Curso'].iloc[0]
    turno  = dados['Acesso']['df']['Turno'].iloc[0] if 'Turno' in dados['Acesso']['df'].columns else ''

    # Escreve rótulo (col A) e turno (col B) nas células 7-12 de todas as abas
    sheets = ['Acesso', 'Matrícula', 'Conclusão', 'Evasão', 'Retenção', 'Eficiência', 'PAP']
    for sheet in sheets:
        for linha in range(7, 13):
            wb[sheet].cell(row=linha, column=1, value=rotulo)
            wb[sheet].cell(row=linha, column=2, value=turno)

    # Escreve cada tabela na aba correspondente usando a lista (col, linha, df)
    mapa_abas = {
        'Matrícula': dados['Matrícula']['tabelas'],
        'Conclusão': dados['Conclusão']['tabelas'],
        'Evasão':    dados['Evasão']['tabelas'],
        'Retenção':  dados['Retenção']['tabelas'],
        'Eficiência': dados['Eficiência']['tabelas'],
    }
    for aba, tabelas in mapa_abas.items():
        for col, linha, df in tabelas:
            escreve_tabela(wb, df, aba, col, linha)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
